from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import pandas as pd


def write_to_excel(output, data):
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        # Access the workbook
        workbook = writer.book
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Regular font
        garamond_font = Font(name='Garamond', size=11)
        # Header font (bold)
        header_font = Font(name='Garamond', size=12, bold=True)
        # Header background color (optional)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        # Center alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        # Apply formatting to each sheet
        for sheet_name in data.keys():
            worksheet = workbook[sheet_name]
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            # Format all cells
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    # Header row formatting
                    if row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    else:
                        cell.font = garamond_font
            # Auto-adjust column widths (optional)
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
